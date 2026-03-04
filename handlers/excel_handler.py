from datetime import datetime, date

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

from utils.security import validate_image_path
from utils.file_utils import generate_output_path

HEADER_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
HEADER_FONT = Font(bold=True, size=10, name="Calibri")
CELL_FONT = Font(size=10, name="Calibri")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def create_excel(
    title: str,
    sheets: list[dict],
    output_path: str = "",
) -> str:
    """Generate an Excel workbook from structured sheet definitions.

    Each sheet dict:
      {
        "name": "Sheet1",
        "headers": ["Col A", "Col B"],
        "rows": [["v1", "v2"], ...],
        "column_widths": [20, 30],           # optional
        "images": [{"path": "...", "cell": "E1"}],  # optional
      }

    Returns the absolute path of the generated file.
    """
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_def in sheets:
        sheet_name = sheet_def.get("name", "Sheet")
        ws = wb.create_sheet(title=sheet_name)
        headers = sheet_def.get("headers", [])
        rows = sheet_def.get("rows", [])
        col_widths = sheet_def.get("column_widths", [])
        images = sheet_def.get("images", [])

        for j, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=j, value=str(h))
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

        col_types = sheet_def.get("column_types", [])
        date_fmt = sheet_def.get("date_format", "DD/MM/YYYY")

        for i, row in enumerate(rows, 2):
            for j, val in enumerate(row, 1):
                col_type = col_types[j - 1] if j - 1 < len(col_types) else "auto"
                converted = _typed_value(val, col_type)
                cell = ws.cell(row=i, column=j, value=converted)
                cell.font = CELL_FONT
                cell.border = THIN_BORDER
                if isinstance(converted, (datetime, date)):
                    cell.number_format = _openpyxl_date_format(date_fmt)

        for j, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w

        if not col_widths and headers:
            for j, h in enumerate(headers, 1):
                max_len = len(str(h))
                for row in rows:
                    if j - 1 < len(row):
                        max_len = max(max_len, len(str(row[j - 1])))
                ws.column_dimensions[get_column_letter(j)].width = min(max_len + 4, 50)

        for img_def in images:
            img_path = validate_image_path(img_def["path"])
            xl_img = XlImage(str(img_path))
            ws.add_image(xl_img, img_def.get("cell", "A1"))

    dest = generate_output_path(output_path)
    wb.save(str(dest))
    return str(dest)


def read_excel(file_path: str, sheet_name: str | None = None) -> dict:
    """Extract data from an Excel workbook.

    Args:
        file_path: Path to the .xlsx file.
        sheet_name: Specific sheet to read. None = all sheets.

    Returns:
        {
            "sheets": [
                {
                    "name": str,
                    "headers": [...],
                    "rows": [[...]],
                    "row_count": int,
                    "column_count": int
                }
            ],
            "sheet_names": [...]
        }
    """
    resolved = validate_path(file_path, must_exist=True)
    wb = load_workbook(str(resolved), read_only=True, data_only=True)

    target_sheets = [sheet_name] if sheet_name else wb.sheetnames
    sheets_data = []

    for sn in target_sheets:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        rows_raw = list(ws.iter_rows(values_only=True))
        if not rows_raw:
            sheets_data.append({
                "name": sn, "headers": [], "rows": [],
                "row_count": 0, "column_count": 0,
            })
            continue

        headers = [str(c) if c is not None else "" for c in rows_raw[0]]
        rows = []
        for row in rows_raw[1:]:
            rows.append([str(c) if c is not None else "" for c in row])

        sheets_data.append({
            "name": sn,
            "headers": headers,
            "rows": rows,
            "row_count": len(rows),
            "column_count": len(headers),
        })

    wb.close()
    return {"sheets": sheets_data, "sheet_names": wb.sheetnames}


_DATE_FORMATS = [
    ("%d/%m/%Y", "DD/MM/YYYY"),
    ("%Y-%m-%d", "YYYY-MM-DD"),
    ("%d-%m-%Y", "DD-MM-YYYY"),
    ("%m/%d/%Y", "MM/DD/YYYY"),
    ("%Y/%m/%d", "YYYY/MM/DD"),
    ("%d.%m.%Y", "DD.MM.YYYY"),
]

_OPENPYXL_DATE_FMT_MAP = {
    "DD/MM/YYYY": "DD/MM/YYYY",
    "YYYY-MM-DD": "YYYY-MM-DD",
    "DD-MM-YYYY": "DD-MM-YYYY",
    "MM/DD/YYYY": "MM/DD/YYYY",
    "YYYY/MM/DD": "YYYY/MM/DD",
    "DD.MM.YYYY": "DD.MM.YYYY",
}


def _openpyxl_date_format(fmt: str) -> str:
    return _OPENPYXL_DATE_FMT_MAP.get(fmt, "DD/MM/YYYY")


def _parse_date(val) -> datetime | None:
    if isinstance(val, (datetime, date)):
        return val if isinstance(val, datetime) else datetime(val.year, val.month, val.day)
    s = str(val).strip()
    for py_fmt, _ in _DATE_FORMATS:
        try:
            return datetime.strptime(s, py_fmt)
        except ValueError:
            continue
    return None


def _smart_value(val):
    """Attempt numeric conversion for cell values."""
    if val is None:
        return ""
    if isinstance(val, (int, float)):
        return val
    s = str(val).strip()
    try:
        return int(s)
    except ValueError:
        pass
    try:
        return float(s)
    except ValueError:
        pass
    return s


def _typed_value(val, col_type: str):
    """Convert a cell value according to the declared column type.

    Supported col_type values:
        "string"  - always store as text
        "number"  - force numeric (int or float)
        "date"    - parse into a datetime object for native Excel date cell
        "auto"    - use _smart_value heuristic (default)
    """
    if val is None or (isinstance(val, str) and val.strip() == ""):
        return ""

    if col_type == "string":
        return str(val)

    if col_type == "number":
        if isinstance(val, (int, float)):
            return val
        s = str(val).strip()
        try:
            return int(s)
        except ValueError:
            pass
        try:
            return float(s)
        except ValueError:
            return s

    if col_type == "date":
        parsed = _parse_date(val)
        return parsed if parsed is not None else str(val)

    return _smart_value(val)
